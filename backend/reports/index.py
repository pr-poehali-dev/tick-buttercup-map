import json
import os
import psycopg2
import requests
from datetime import datetime, timedelta
from io import BytesIO
from typing import Dict, Any

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    '''
    Business: Автоматическая генерация и отправка отчетов по меткам
    Args: event - HTTP запрос (можно вызывать по расписанию)
          context - контекст выполнения
    Returns: Результат отправки отчета
    '''
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, X-Admin-Token',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    database_url = os.environ.get('DATABASE_URL')
    bot_token = os.environ.get('TELEGRAM_BOT_TOKEN')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID')
    
    headers = {
        'Content-Type': 'application/json',
        'Access-Control-Allow-Origin': '*'
    }
    
    try:
        conn = psycopg2.connect(database_url)
        cursor = conn.cursor()
        
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)
        
        cursor.execute('''
            SELECT 
                id, type, latitude, longitude, 
                TO_CHAR(created_at, 'DD.MM.YYYY HH24:MI'), 
                description, verified
            FROM marks
            WHERE DATE(created_at) = %s
            ORDER BY created_at DESC
        ''', (yesterday,))
        
        marks_data = cursor.fetchall()
        
        cursor.execute('''
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN type = 'tick' THEN 1 ELSE 0 END) as tick_count,
                SUM(CASE WHEN type = 'hogweed' THEN 1 ELSE 0 END) as hogweed_count,
                SUM(CASE WHEN verified = true THEN 1 ELSE 0 END) as verified_count
            FROM marks
            WHERE DATE(created_at) = %s
        ''', (yesterday,))
        
        stats = cursor.fetchone()
        total_marks = stats[0] or 0
        tick_count = stats[1] or 0
        hogweed_count = stats[2] or 0
        verified_count = stats[3] or 0
        
        cursor.close()
        conn.close()
        
        if total_marks == 0:
            return {
                'statusCode': 200,
                'headers': headers,
                'body': json.dumps({'message': 'Нет меток за вчера'}),
                'isBase64Encoded': False
            }
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Отчет {yesterday.strftime('%d.%m.%Y')}"
            
            ws['A1'] = 'ОТЧЕТ ПО МЕТКАМ КЛЕЩЕЙ И БОРЩЕВИКА'
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:G1')
            
            ws['A2'] = f"Дата: {yesterday.strftime('%d.%m.%Y')}"
            ws['A2'].font = Font(bold=True)
            
            ws['A4'] = 'СТАТИСТИКА'
            ws['A4'].font = Font(size=12, bold=True)
            ws['A5'] = f'Всего меток: {total_marks}'
            ws['A6'] = f'Клещи: {tick_count}'
            ws['A7'] = f'Борщевик: {hogweed_count}'
            ws['A8'] = f'Проверено: {verified_count}'
            
            header_row = 10
            headers_list = ['ID', 'Тип', 'Широта', 'Долгота', 'Дата/Время', 'Описание', 'Статус']
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, header in enumerate(headers_list, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for row_num, mark in enumerate(marks_data, header_row + 1):
                ws.cell(row=row_num, column=1, value=mark[0])
                ws.cell(row=row_num, column=2, value='Клещ' if mark[1] == 'tick' else 'Борщевик')
                ws.cell(row=row_num, column=3, value=float(mark[2]))
                ws.cell(row=row_num, column=4, value=float(mark[3]))
                ws.cell(row=row_num, column=5, value=mark[4])
                ws.cell(row=row_num, column=6, value=mark[5] or '-')
                ws.cell(row=row_num, column=7, value='Проверено' if mark[6] else 'На проверке')
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            if bot_token and chat_id:
                telegram_message = f"""
📊 ЕЖЕДНЕВНЫЙ ОТЧЕТ
📅 Дата: {yesterday.strftime('%d.%m.%Y')}

📍 Всего меток: {total_marks}
🦟 Клещи: {tick_count}
🌿 Борщевик: {hogweed_count}
✅ Проверено: {verified_count}

Полный отчет во вложении ⬇️
                """
                
                requests.post(
                    f'https://api.telegram.org/bot{bot_token}/sendMessage',
                    json={'chat_id': chat_id, 'text': telegram_message}
                )
                
                requests.post(
                    f'https://api.telegram.org/bot{bot_token}/sendDocument',
                    files={'document': (f'Отчет_{yesterday.strftime("%d.%m.%Y")}.xlsx', excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                    data={'chat_id': chat_id}
                )
        
        except ImportError:
            if bot_token and chat_id:
                text_report = f"""
📊 ЕЖЕДНЕВНЫЙ ОТЧЕТ
📅 Дата: {yesterday.strftime('%d.%m.%Y')}

📍 Всего меток: {total_marks}
🦟 Клещи: {tick_count}
🌿 Борщевик: {hogweed_count}
✅ Проверено: {verified_count}

СПИСОК МЕТОК:
"""
                for mark in marks_data[:10]:
                    mark_type = 'Клещ' if mark[1] == 'tick' else 'Борщевик'
                    text_report += f"\n• {mark_type} ({mark[2]:.4f}, {mark[3]:.4f}) - {mark[4]}"
                
                if len(marks_data) > 10:
                    text_report += f"\n\n...и еще {len(marks_data) - 10} меток"
                
                requests.post(
                    f'https://api.telegram.org/bot{bot_token}/sendMessage',
                    json={'chat_id': chat_id, 'text': text_report}
                )
        
        return {
            'statusCode': 200,
            'headers': headers,
            'body': json.dumps({'success': True, 'marks_count': total_marks}),
            'isBase64Encoded': False
        }
    
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': headers,
            'body': json.dumps({'error': str(e)}),
            'isBase64Encoded': False
        }
